import numpy    as np
import openpyxl as oxl

##sdex = 1
##wb_all    = oxl.load_workbook(filename='MUestimates_all_locations_1.xlsx')
##wb_home   = oxl.load_workbook(filename='MUestimates_home_1.xlsx')
##wb_work   = oxl.load_workbook(filename='MUestimates_work_1.xlsx')
##wb_school = oxl.load_workbook(filename='MUestimates_school_1.xlsx')
##wb_other  = oxl.load_workbook(filename='MUestimates_other_locations_1.xlsx')

sdex = 0
wb_all    = oxl.load_workbook(filename='MUestimates_all_locations_2.xlsx')
wb_home   = oxl.load_workbook(filename='MUestimates_home_2.xlsx')
wb_work   = oxl.load_workbook(filename='MUestimates_work_2.xlsx')
wb_school = oxl.load_workbook(filename='MUestimates_school_2.xlsx')
wb_other  = oxl.load_workbook(filename='MUestimates_other_locations_2.xlsx')

wsname  = 'South Africa'


ws01 = wb_all[wsname]
ws02 = wb_home[wsname]
ws03 = wb_work[wsname]
ws04 = wb_school[wsname]
ws05 = wb_other[wsname]

val01 = list()
val02 = list()
val03 = list()
val04 = list()
val05 = list()

for rowval in ws01.rows:
  val01.append([rowval[k1].value for k1 in range(16)])
for rowval in ws02.rows:
  val02.append([rowval[k1].value for k1 in range(16)])
for rowval in ws03.rows:
  val03.append([rowval[k1].value for k1 in range(16)])
for rowval in ws04.rows:
  val04.append([rowval[k1].value for k1 in range(16)])
for rowval in ws05.rows:
  val05.append([rowval[k1].value for k1 in range(16)])

mat01 = np.array(val01[sdex:])
mat02 = np.array(val02[sdex:])
mat03 = np.array(val03[sdex:])
mat04 = np.array(val04[sdex:])
mat05 = np.array(val05[sdex:])

rowform  = '     ['+15*'{:4.2e},'+'{:4.2e}],\n'

mat02list = (mat02).tolist()
with open('mat_home.txt','w') as fid01:
  for rowval in mat02list:
    fid01.write(rowform.format(*rowval))

mat03list = (mat03).tolist()
with open('mat_work.txt','w') as fid01:
  for rowval in mat03list:
    fid01.write(rowform.format(*rowval))

mat04list = (mat04).tolist()
with open('mat_schl.txt','w') as fid01:
  for rowval in mat04list:
    fid01.write(rowform.format(*rowval))

mat05list = (mat05).tolist()
with open('mat_comm.txt','w') as fid01:
  for rowval in mat05list:
    fid01.write(rowform.format(*rowval))


##age_pyr = [0.085,0.085,0.091,0.091,0.089,0.085,0.081,0.075,
##           0.065,0.058,0.050,0.043,0.035,0.028,0.018,0.021]
##ages = np.array(age_pyr)
##
##R0val1 = np.dot(np.dot(ages,mat01),ages)
##
##dcv02 = np.round(np.dot(np.dot(ages,mat02/R0val1),ages),3)
##dcv03 = np.round(np.dot(np.dot(ages,mat03/R0val1),ages),3)
##dcv04 = np.round(np.dot(np.dot(ages,mat04/R0val1),ages),3)
##dcv05 = np.round(np.dot(np.dot(ages,mat05/R0val1),ages),3)
##                 
##print(dcv02,dcv03,dcv04,dcv05,dcv02+dcv03+dcv04+dcv05)


